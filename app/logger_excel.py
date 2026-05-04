from datetime import datetime
from openpyxl import Workbook, load_workbook
from app.config import TRADE_LOG_FILE


TRADE_HEADERS = [
    "TradeID",
    "EntryTime",
    "ExitTime",
    "Symbol",
    "Instrument",
    "Side",
    "Qty",
    "Entry",
    "Exit",
    "PnL",
    "Result",
    "Reason",
    "TradeCount",
    "Status",
    "Equity",
    "ATR",
    "SL",
    "Target",
    "EntryType",
    "DurationSec",
]

METRIC_HEADERS = ["Metric", "Value"]


def init_excel():
    if TRADE_LOG_FILE.exists():
        return

    wb = Workbook()

    ws = wb.active
    ws.title = "Trades"
    ws.append(TRADE_HEADERS)

    ws2 = wb.create_sheet("Metrics")
    ws2.append(METRIC_HEADERS)

    wb.save(TRADE_LOG_FILE)


def _safe_save_workbook(wb):
    try:
        wb.save(TRADE_LOG_FILE)
        return True
    except PermissionError:
        print("⚠️ Close Excel file before logging trades!")
        return False


def _ensure_headers():
    init_excel()

    wb = load_workbook(TRADE_LOG_FILE)

    if "Trades" not in wb.sheetnames:
        ws = wb.create_sheet("Trades", 0)
        ws.append(TRADE_HEADERS)
    else:
        ws = wb["Trades"]
        if ws.max_row == 0:
            ws.append(TRADE_HEADERS)
        elif ws.max_row >= 1:
            existing = [cell.value for cell in ws[1]]
            if existing != TRADE_HEADERS:
                ws.delete_rows(1, ws.max_row)
                ws.append(TRADE_HEADERS)

    if "Metrics" not in wb.sheetnames:
        ws2 = wb.create_sheet("Metrics")
        ws2.append(METRIC_HEADERS)
    else:
        ws2 = wb["Metrics"]
        if ws2.max_row == 0:
            ws2.append(METRIC_HEADERS)
        elif ws2.max_row >= 1:
            existing = [cell.value for cell in ws2[1]]
            if existing != METRIC_HEADERS:
                ws2.delete_rows(1, ws2.max_row)
                ws2.append(METRIC_HEADERS)

    _safe_save_workbook(wb)
    wb.close()


def log_trade_entry(
    trade_id,
    symbol,
    instrument,
    side,
    qty,
    entry_price,
    trade_count,
    reason="Trade opened",
    atr=None,
    sl=None,
    target=None,
    entry_type=None,
):
    _ensure_headers()

    wb = load_workbook(TRADE_LOG_FILE)
    ws = wb["Trades"]

    # Duplicate protection
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(trade_id):
            print("⚠️ Duplicate trade entry prevented")
            wb.close()
            return

    ws.append([
        trade_id,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        None,
        symbol,
        instrument,
        side,
        qty,
        entry_price,
        None,
        None,
        None,
        reason,
        trade_count,
        "OPEN",
        None,
        atr,
        sl,
        target,
        entry_type,
        None,
    ])

    _safe_save_workbook(wb)
    wb.close()


def log_trade_exit(trade_id, exit_price, result, reason, status="CLOSED"):
    _ensure_headers()

    wb = load_workbook(TRADE_LOG_FILE)
    ws = wb["Trades"]

    for row in ws.iter_rows(min_row=2):
        cell_trade_id = row[0].value

        if str(cell_trade_id) != str(trade_id):
            continue

        entry_time_str = row[1].value
        exit_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row[2].value = exit_time_str
        row[8].value = exit_price

        entry_price = float(row[7].value) if row[7].value is not None else 0.0
        qty = float(row[6].value) if row[6].value is not None else 0.0
        side = str(row[5].value).upper() if row[5].value else "BUY"

        if side == "BUY":
            pnl = (float(exit_price) - entry_price) * qty
        else:
            pnl = (entry_price - float(exit_price)) * qty

        row[9].value = round(pnl, 2)
        row[10].value = result
        row[11].value = reason
        row[13].value = status

        try:
            entry_dt = datetime.strptime(str(entry_time_str), "%Y-%m-%d %H:%M:%S")
            exit_dt = datetime.strptime(exit_time_str, "%Y-%m-%d %H:%M:%S")
            duration_sec = int((exit_dt - entry_dt).total_seconds())
        except Exception:
            duration_sec = None

        row[19].value = duration_sec
        break

    saved = _safe_save_workbook(wb)
    wb.close()

    if saved:
        update_metrics()


def update_metrics():
    _ensure_headers()

    wb = load_workbook(TRADE_LOG_FILE)
    ws = wb["Trades"]
    metrics = wb["Metrics"]

    pnls = []
    equity_curve = []
    running_total = 0.0

    for row in ws.iter_rows(min_row=2):
        pnl_value = row[9].value

        if pnl_value is None:
            continue

        try:
            pnl = float(pnl_value)
        except Exception:
            continue

        pnls.append(pnl)
        running_total += pnl
        equity_curve.append(running_total)
        row[14].value = round(running_total, 2)

    if metrics.max_row > 1:
        metrics.delete_rows(2, metrics.max_row - 1)

    if not pnls:
        _safe_save_workbook(wb)
        wb.close()
        return

    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]

    win_rate = (len(wins) / len(pnls)) * 100 if pnls else 0.0
    avg_win = sum(wins) / len(wins) if wins else 0.0
    avg_loss = sum(losses) / len(losses) if losses else 0.0
    profit_factor = abs(sum(wins) / sum(losses)) if losses else float("inf")

    peak = equity_curve[0] if equity_curve else 0.0
    max_drawdown = 0.0

    for value in equity_curve:
        if value > peak:
            peak = value
        drawdown = peak - value
        if drawdown > max_drawdown:
            max_drawdown = drawdown

    data = [
        ("Total Trades", len(pnls)),
        ("Wins", len(wins)),
        ("Losses", len(losses)),
        ("Win Rate %", round(win_rate, 2)),
        ("Net PnL", round(running_total, 2)),
        ("Avg Win", round(avg_win, 2)),
        ("Avg Loss", round(avg_loss, 2)),
        ("Profit Factor", round(profit_factor, 2) if profit_factor != float("inf") else "inf"),
        ("Max Drawdown", round(max_drawdown, 2)),
    ]

    for item in data:
        metrics.append(item)

    _safe_save_workbook(wb)
    wb.close()
import json
from datetime import datetime
import pandas as pd

from app.utils import _deep_deserialize, _serialize_for_json
from app.config import (
    BOT_CONTROL_FILE,
    BOT_STATE_FILE,
    MARKET_DATA_FILE,
    MAX_TRADES,
    START_MODE,
    PAPER_TRADING,
    RECONCILE_WITH_BROKER_ON_STARTUP,
    TICK_LOG_FILE,
)
from app.state import STATE


MARKET_DATA_COLUMNS = [
    "time", "open", "high", "low", "close", "volume",
    "ema20", "ema50", "vwap", "signal"
]

TICK_DATA_COLUMNS = ["time", "ltp", "ltq"]

#------------------------------------------------------------------

# dropna(...) → remove bad/incomplete rows
#filter → keep only today’s data
#sort_values("time") → arrange in time order
#reset_index → clean index
def read_market_data():
    if not MARKET_DATA_FILE.exists():
        return pd.DataFrame()

    try:
        df = pd.read_csv(MARKET_DATA_FILE)

        if df.empty:
            return df

        if "time" in df.columns:
            df["time"] = pd.to_datetime(
                df["time"],
                errors="coerce",
                format="%Y-%m-%d %H:%M:%S"
            )

        numeric_cols = ["open", "high", "low", "close", "ema20", "ema50", "vwap", "volume"]
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        df = df.dropna(subset=["time", "open", "high", "low", "close"], how="any")
        
        df = df.sort_values("time").reset_index(drop=True)
        return df

    except Exception:
        return pd.DataFrame()




def _parse_mixed_market_time(series):
    s = series.astype(str).str.strip()

    def parse_one(x):
        if not x or x.lower() == "nan":
            return pd.NaT

        # local CSV style
        for fmt in ("%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return pd.Timestamp(datetime.strptime(x, fmt))
            except Exception:
                pass

        # fallback for API style like 2026-04-16T11:30:00+05:30
        try:
            ts = pd.to_datetime(x, errors="coerce")
            if pd.isna(ts):
                return pd.NaT

            # make timezone-aware values naive in local time
            if getattr(ts, "tzinfo", None) is not None:
                ts = ts.tz_localize(None)

            return pd.Timestamp(ts)
        except Exception:
            return pd.NaT

    return s.apply(parse_one)

def _safe_float(value, default=None):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _empty_market_df():
    return pd.DataFrame(columns=MARKET_DATA_COLUMNS)


def _normalize_market_df(df):
    if df is None or df.empty:
        return _empty_market_df()

    df = df.copy()

    for col in MARKET_DATA_COLUMNS:
        if col not in df.columns:
            df[col] = None

    df = df[MARKET_DATA_COLUMNS]

    df["time"] = _parse_mixed_market_time(df["time"])
    df = df.dropna(subset=["time"])

    for col in ["open", "high", "low", "close", "ema20", "ema50", "vwap"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df["volume"] = pd.to_numeric(df["volume"], errors="coerce").fillna(0.0)
    df["signal"] = df["signal"].fillna("").astype(str)

    df = df.dropna(subset=["open", "high", "low", "close"])
    df = df.sort_values("time").drop_duplicates(subset=["time"], keep="last")

    return df


def _read_market_df():
    ensure_market_file()

    try:
        df = pd.read_csv(MARKET_DATA_FILE)
        normalized = _normalize_market_df(df)
        # If main file is empty/corrupt but backup exists, restore from backup
        if normalized.empty:
            from pathlib import Path
            bak_file = Path(MARKET_DATA_FILE).with_suffix(".csv.bak")
            if bak_file.exists() and bak_file.stat().st_size > 50:
                print("⚠️ Main CSV empty — restoring from backup")
                df_bak = pd.read_csv(bak_file)
                normalized = _normalize_market_df(df_bak)
                if not normalized.empty:
                    _write_market_df(normalized)
                    return normalized
        return normalized
    except Exception:
        # Try backup on read failure
        from pathlib import Path
        bak_file = Path(MARKET_DATA_FILE).with_suffix(".csv.bak")
        if bak_file.exists() and bak_file.stat().st_size > 50:
            try:
                print("⚠️ Main CSV corrupt — restoring from backup")
                df_bak = pd.read_csv(bak_file)
                normalized = _normalize_market_df(df_bak)
                if not normalized.empty:
                    _write_market_df(normalized)
                    return normalized
            except Exception:
                pass
        return _empty_market_df()


def _write_market_df(df):
    df = _normalize_market_df(df)

    if not df.empty:
        df = df.copy()
        df["time"] = df["time"].dt.strftime("%Y-%m-%d %H:%M:%S")

    # Atomic write: write to temp file, then rename to prevent data loss on power failure
    from pathlib import Path
    import shutil

    target = Path(MARKET_DATA_FILE)
    tmp_file = target.with_suffix(".csv.tmp")
    bak_file = target.with_suffix(".csv.bak")

    try:
        df.to_csv(tmp_file, index=False)

        # Verify temp file is valid before replacing
        if tmp_file.stat().st_size > 0:
            # Backup current good file
            if target.exists() and target.stat().st_size > 50:
                shutil.copy2(target, bak_file)
            # Atomic rename
            tmp_file.replace(target)
        else:
            tmp_file.unlink(missing_ok=True)
    except Exception as e:
        print(f"⚠️ Atomic write failed: {e}")
        tmp_file.unlink(missing_ok=True)
        # Fallback: direct write
        df.to_csv(MARKET_DATA_FILE, index=False)

    # Sync to database
    try:
        from app.trade_db import save_market_candles_bulk
        if not df.empty:
            rows = []
            for _, row in df.iterrows():
                rows.append((
                    str(row.get("time", "")),
                    float(row.get("open", 0)),
                    float(row.get("high", 0)),
                    float(row.get("low", 0)),
                    float(row.get("close", 0)),
                    float(row.get("volume", 0) or 0),
                    float(row["ema20"]) if pd.notna(row.get("ema20")) else None,
                    float(row["ema50"]) if pd.notna(row.get("ema50")) else None,
                    float(row["vwap"]) if pd.notna(row.get("vwap")) else None,
                    str(row["signal"]) if pd.notna(row.get("signal")) and str(row.get("signal", "")).strip() else None,
                ))
            save_market_candles_bulk(rows)
    except Exception as e:
        print(f"⚠️ Market data DB sync failed: {e}")


def write_state(status="RUNNING"):
    payload = {
        "ws_connected": getattr(STATE, "ws_connected", False),
        "ltp": getattr(STATE, "ltp", None),
        "trade_count": getattr(STATE, "trade_count", 0),
        "max_trades": MAX_TRADES,
        "current_position": getattr(STATE, "current_position", None),
        "realized_pnl": getattr(STATE, "realized_pnl", 0.0),
        "consecutive_sl": getattr(STATE, "consecutive_sl", 0),
        "last_exit_time": getattr(STATE, "last_exit_time", None),
        "bot_block_reason": getattr(STATE, "bot_block_reason", None),
        "last_update": datetime.now(),
        "status": status,
        "last_signal": getattr(STATE, "last_signal", None),
        "last_action": getattr(STATE, "last_action", None),
        "current_trade_id": getattr(STATE, "current_trade_id", None),
        "pending_trade": getattr(STATE, "pending_trade", None),
        "last_candle_key": getattr(STATE, "last_candle_key", None),
        "last_entry_eval_candle_key": getattr(STATE, "last_entry_eval_candle_key", None),
        "live_candle": getattr(STATE, "current_candle", None),
    }
    
    payload = _serialize_for_json(payload)
    BOT_STATE_FILE.write_text(json.dumps(payload, indent=2, default=str))
    


def ensure_tick_file():
    if not TICK_LOG_FILE.exists():
        pd.DataFrame(columns=TICK_DATA_COLUMNS).to_csv(TICK_LOG_FILE, index=False)


def write_tick_data(ltp, tick_time, ltq=0.0):
    ensure_tick_file()

    row_df = pd.DataFrame([{
        "time": tick_time.strftime("%Y-%m-%d %H:%M:%S.%f"),
        "ltp": float(ltp),
        "ltq": float(ltq),
    }])

    try:
        row_df.to_csv(
            TICK_LOG_FILE,
            mode="a",
            header=not TICK_LOG_FILE.exists() or TICK_LOG_FILE.stat().st_size == 0,
            index=False,
        )
    except Exception as e:
        print(f"write_tick_data failed: {e}")


def load_ticks_after_entry(entry_dt, max_rows=50000):
    ensure_tick_file()

    try:
        df = pd.read_csv(TICK_LOG_FILE)
        if df.empty or "time" not in df.columns:
            return []

        df["time"] = _parse_mixed_market_time(df["time"])
        df["ltp"] = pd.to_numeric(df["ltp"], errors="coerce")
        df = df.dropna(subset=["time", "ltp"])

        if entry_dt is not None:
            df = df[df["time"] > entry_dt]

        if max_rows and len(df) > max_rows:
            df = df.tail(max_rows)

        return df[["time", "ltp"]].values.tolist()

    except Exception as e:
        print(f"load_ticks_after_entry failed: {e}")
        return []


def ensure_control_file():
    if not BOT_CONTROL_FILE.exists():
        BOT_CONTROL_FILE.write_text(json.dumps({"run_bot": True}, indent=2))


def ensure_market_file():
    if not MARKET_DATA_FILE.exists():
        _empty_market_df().to_csv(MARKET_DATA_FILE, index=False)
        return

    try:
        df = pd.read_csv(MARKET_DATA_FILE)
        for col in MARKET_DATA_COLUMNS:
            if col not in df.columns:
                df[col] = None
        df = df[MARKET_DATA_COLUMNS]
        df.to_csv(MARKET_DATA_FILE, index=False)
    except Exception:
        _empty_market_df().to_csv(MARKET_DATA_FILE, index=False)


def should_run_bot():
    ensure_control_file()
    try:
        return json.loads(BOT_CONTROL_FILE.read_text()).get("run_bot", True)
    except Exception:
        return True


def write_control(run_bot: bool):
    BOT_CONTROL_FILE.write_text(json.dumps({"run_bot": run_bot}, indent=2))


def restore_candles_from_market_data(history_limit=500):
    """
    Restore recent candle history into runtime memory.

    Purpose:
    - Keep previous-session candles in memory for EMA / ATR / signal warm-up
    - Dashboard will still show only today's candles because dashboard.py
      already filters MARKET_DATA_FILE rows by today's date
    """
    try:
        df = _read_market_df()

        if df.empty:
            print("No market history available to restore")
            with STATE.lock:
                STATE.latest_closed_candles = []
            return 0

        df = df.sort_values("time").reset_index(drop=True)

        if history_limit and len(df) > history_limit:
            df = df.tail(history_limit).copy()

        restored = []
        for _, row in df.iterrows():
            restored.append([
                row["time"].strftime("%Y-%m-%d %H:%M:%S"),
                float(row["open"]),
                float(row["high"]),
                float(row["low"]),
                float(row["close"]),
                float(row["volume"]),
            ])

        with STATE.lock:
            STATE.latest_closed_candles = restored

        today = pd.Timestamp.now().date()
        today_count = int((df["time"].dt.date == today).sum())
        warmup_count = len(restored) - today_count

        print(f"📦 Warm-up candles in memory: {warmup_count}")
        print(f"📊 Today candles in memory: {today_count}")
        print(f"📊 Restored candles: {len(restored)}")
        return len(restored)

    except Exception as e:
        print(f"restore_candles_from_market_data failed: {e}")
        return 0

def reset_runtime_state():
    with STATE.lock:
        STATE.realized_pnl = 0.0
        STATE.consecutive_sl = 0
        STATE.bot_block_reason = None
        STATE.last_exit_time = None
        STATE.current_position = None
        #STATE.trade_count = 0
        STATE.last_candle_key = None
        STATE.last_entry_eval_candle_key = None
        STATE.pending_trade = None
        STATE.current_trade_id = None
    
        if STATE.last_trade_day != datetime.now().date():
            STATE.trade_count = 0
    write_state("CLOSED")
    print("🔁 State reset -> CLOSED")


def _restore_from_data(data):
    
    if PAPER_TRADING and not BOT_STATE_FILE.exists():
        reset_runtime_state()
        return
    data = _deep_deserialize(data)
    with STATE.lock:
        STATE.realized_pnl = data.get("realized_pnl", 0.0)
        STATE.consecutive_sl = data.get("consecutive_sl", 0)
        STATE.bot_block_reason = data.get("bot_block_reason")
        STATE.last_exit_time = data.get("last_exit_time")
        STATE.current_position = data.get("current_position")
        STATE.trade_count = data.get("trade_count", 0)
        STATE.last_candle_key = data.get("last_candle_key")
        STATE.last_entry_eval_candle_key = data.get("last_entry_eval_candle_key")
        STATE.pending_trade = data.get("pending_trade")
        STATE.current_trade_id = data.get("current_trade_id")
    
    
    print("✅ State restored successfully")


def restore_runtime_state():
    mode = START_MODE.lower().strip()

    if not BOT_STATE_FILE.exists():
        print("No previous state -> fresh start")
        reset_runtime_state()
        with STATE.lock:
            STATE.current_position = None
            STATE.pending_trade = None
            STATE.current_trade_id = None
        return

    try:
        data = json.loads(BOT_STATE_FILE.read_text())
        data = _deep_deserialize(data)
    except Exception as e:
        print(f"State read failed: {e}")
        reset_runtime_state()
        return

    today = datetime.now().date()
    last_update = data.get("last_update")
    last_date = None

    if last_update:
        try:
            if isinstance(last_update, datetime):
                last_date = last_update.date()
            else:
                last_date = datetime.fromisoformat(str(last_update)).date()
        except Exception:
            last_date = None

    has_open_position_json = data.get("current_position") is not None

    if mode == "fresh":
        print("START_MODE = FRESH -> clearing state")
        reset_runtime_state()
        return

    if mode == "resume":
        print("START_MODE = RESUME -> restoring previous bot state")
        _restore_from_data(data)
        return

    if mode == "auto":
        print("START_MODE = AUTO")

        if PAPER_TRADING:
            if last_date == today:
                print("♻️ Same day (paper) -> resume state")

                rebuilt = rebuild_state_from_trade_log()

                json_trade_count = int(data.get("trade_count", 0) or 0)
                rebuilt_trade_count = int(rebuilt["trade_count"]) if rebuilt else 0

                if rebuilt and rebuilt_trade_count > json_trade_count:
                    print("⚠️ Trade log has higher count than JSON → rebuilding from trade log")
                    with STATE.lock:
                        STATE.trade_count = rebuilt["trade_count"]
                        STATE.realized_pnl = rebuilt["realized_pnl"]
                        STATE.current_position = rebuilt["current_position"]
                        STATE.last_exit_time = rebuilt["last_exit_time"]
                        STATE.last_trade_day = today

                    write_state("RESTORED_FROM_TRADE_LOG")
                    print("✅ State rebuilt from trade log")
                    return

                _restore_from_data(data)
                with STATE.lock:
                    if getattr(STATE, "last_trade_day", None) is None:
                        STATE.last_trade_day = today
                return

            print("🆕 New day (paper) -> fresh start")
            reset_runtime_state()
            with STATE.lock:
                STATE.last_trade_day = today
            return

        if has_open_position_json:
            print("⚠️ Open position found in JSON -> resuming")
            _restore_from_data(data)
            return

        if last_date != today:
            print("🆕 New day detected -> fresh start")
            reset_runtime_state()
            with STATE.lock:
                STATE.last_trade_day = today
            return

        print("♻️ Same day -> resume")
        _restore_from_data(data)
        return

    if mode == "broker_sync":
        print("START_MODE = BROKER_SYNC")
        broker_position = _extract_broker_open_position()

        if broker_position:
            with STATE.lock:
                STATE.current_position = broker_position
                STATE.pending_trade = None
                STATE.current_trade_id = None
                STATE.last_exit_time = None
                STATE.bot_block_reason = None
                STATE.realized_pnl = data.get("realized_pnl", 0.0)
                STATE.consecutive_sl = data.get("consecutive_sl", 0)
                STATE.trade_count = data.get("trade_count", 0)
                STATE.last_candle_key = data.get("last_candle_key")
                STATE.last_entry_eval_candle_key = data.get("last_entry_eval_candle_key")
                STATE.last_trade_day = today

            write_state("BROKER_SYNC_OPEN")
            print("✅ Runtime state rebuilt from broker position")
            return

        if PAPER_TRADING:
            print("Paper mode broker_sync fallback -> using auto logic")
            if last_date == today:
                _restore_from_data(data)
                with STATE.lock:
                    if getattr(STATE, "last_trade_day", None) is None:
                        STATE.last_trade_day = today
            else:
                reset_runtime_state()
                with STATE.lock:
                    STATE.last_trade_day = today
            return

        print("No live broker position -> clearing stale JSON state")
        reset_runtime_state()
        with STATE.lock:
            STATE.last_trade_day = today
        return

    print(f"Unknown START_MODE={START_MODE} -> defaulting to fresh")
    reset_runtime_state()
    with STATE.lock:
        STATE.last_trade_day = today


def _extract_broker_open_position():
    if PAPER_TRADING:
        print("Paper trading mode -> broker sync skipped")
        return None

    if not RECONCILE_WITH_BROKER_ON_STARTUP:
        print("Broker reconciliation disabled")
        return None

    if getattr(STATE, "obj", None) is None:
        print("Broker object not available at startup")
        return None

    try:
        response = STATE.obj.position()

        if not response:
            print("Broker position response empty")
            return None

        data = response.get("data", response) if isinstance(response, dict) else response

        if not isinstance(data, list):
            print("Broker position format unexpected")
            return None

        for pos in data:
            if not isinstance(pos, dict):
                continue

            net_qty = (
                pos.get("netqty")
                or pos.get("netQty")
                or pos.get("buyqty")
                or pos.get("net_quantity")
                or 0
            )

            try:
                net_qty = int(float(net_qty))
            except Exception:
                net_qty = 0

            if net_qty == 0:
                continue

            symbol = pos.get("tradingsymbol") or pos.get("symbolname") or pos.get("symbol")
            token = pos.get("symboltoken") or pos.get("token") or pos.get("symbolToken")
            exchange = pos.get("exchange") or pos.get("exchangeSegment") or "NFO"

            avg_price = (
                pos.get("averageprice")
                or pos.get("avgnetprice")
                or pos.get("avgPrice")
                or pos.get("pnl_avg_price")
                or 0
            )

            try:
                avg_price = float(avg_price)
            except Exception:
                avg_price = 0.0

            side = "BUY" if net_qty > 0 else "SELL"

            instrument = "OPT"
            sym_upper = str(symbol).upper() if symbol else ""
            if "CE" in sym_upper:
                instrument = "CE"
            elif "PE" in sym_upper:
                instrument = "PE"
            elif "FUT" in sym_upper:
                instrument = "FUT"

            normalized = {
                "symbol": symbol,
                "token": token,
                "instrument": instrument,
                "exchange": exchange,
                "side": side,
                "entry_price": avg_price,
                "entry_index_ltp": None,
                "signal_price": avg_price,
                "desired_entry": avg_price,
                "sl": None,
                "target": None,
                "last_trail_price": avg_price,
                "highest_price": avg_price,
                "lowest_price": avg_price,
                "partial_booked": False,
                "partial_qty": 0,
                "remaining_qty": abs(net_qty),
                "status": "OPEN",
                "entry_order_id": "BROKER_SYNC",
                "entry_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "paper_trade": False,
                "entry_logged": True,
                "trade_id": None,
                "entry_reason": "BROKER_SYNC",
            }

            print(f"✅ Broker open position found: {symbol} qty={abs(net_qty)} side={side}")
            return normalized

        print("No open broker position found")
        return None

    except Exception as e:
        print(f"Broker sync failed: {e}")
        return None


def write_market_data(signal=""):
    if getattr(STATE, "shutting_down", False):
        return

    candles = getattr(STATE, "latest_closed_candles", [])
    if not candles:
        return

    last_candle = candles[-1]
    candle_time = str(last_candle[0])

    instrument_label = ""
    if getattr(STATE, "current_position", None):
        sym = str(STATE.current_position.get("symbol", "")).upper()
        if "CE" in sym:
            instrument_label = "CE"
        elif "PE" in sym:
            instrument_label = "PE"
        elif "FUT" in sym:
            instrument_label = "FUT"

    final_signal = signal.strip() if signal else ""
    if final_signal and instrument_label:
        if final_signal == "BUY":
            final_signal = f"{instrument_label} BUY"
        elif final_signal == "SELL":
            final_signal = f"{instrument_label} SELL"
        elif final_signal == "PARTIAL":
            final_signal = f"{instrument_label} PARTIAL"

    row = {
        "time": candle_time,
        "open": _safe_float(last_candle[1]),
        "high": _safe_float(last_candle[2]),
        "low": _safe_float(last_candle[3]),
        "close": _safe_float(last_candle[4]),
        "volume": _safe_float(last_candle[5], 0.0) if len(last_candle) > 5 else 0.0,
        "ema20": _safe_float(getattr(STATE, "ema20", None)),
        "ema50": _safe_float(getattr(STATE, "ema50", None)),
        "vwap": _safe_float(getattr(STATE, "vwap", None)),
        "signal": final_signal if final_signal else "",
    }

    try:
        df = _read_market_df()

        if not df.empty and candle_time in df["time"].dt.strftime("%Y-%m-%d %H:%M:%S").values:
            idx = df.index[df["time"].dt.strftime("%Y-%m-%d %H:%M:%S") == candle_time][-1]

            existing_signal = str(df.at[idx, "signal"]).strip().upper()
            new_signal = str(final_signal).strip().upper()

            for col in row:
                if col != "signal":
                    df.at[idx, col] = row[col]
                else:
                    if not new_signal:
                        continue
                    existing_is_labeled = existing_signal in (
                        "CE BUY", "CE SELL", "PE BUY", "PE SELL", "FUT BUY", "FUT SELL"
                    )
                    new_is_generic = new_signal in ("BUY", "SELL", "PARTIAL")
                    if existing_is_labeled and new_is_generic:
                        continue
                    df.at[idx, col] = row[col]
        else:
            new_row_df = pd.DataFrame([row], columns=MARKET_DATA_COLUMNS)
            df = pd.concat([df, new_row_df], ignore_index=True)

        _write_market_df(df)

    except Exception as e:
        print(f"write_market_data failed: {e}")


def get_last_saved_candle_time():
    try:
        df = _read_market_df()
        if df.empty:
            return None
        return df["time"].max()
    except Exception as e:
        print(f"❌ Error reading last candle time: {e}")
        return None


def overwrite_market_data_with_candles(candles):
    """
    Merge new candles into existing market data.
    Keeps existing rows, adds new ones, deduplicates by time.
    """
    if not candles:
        return 0

    try:
        df_existing = _read_market_df()

        df_new = pd.DataFrame(candles, columns=[
            "time", "open", "high", "low", "close", "volume"
        ])

        for col in ["ema20", "ema50", "vwap", "signal"]:
            if col not in df_new.columns:
                df_new[col] = None if col != "signal" else ""

        df_new = df_new[MARKET_DATA_COLUMNS]

        if df_existing is not None and not df_existing.empty:
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            df_combined = df_new.copy()

        _write_market_df(df_combined)
        from app.indicators import recompute_indicators_full
        recompute_indicators_full()
        final_df = _read_market_df()
        print(f"✅ Overwrite rebuild complete | candles in file: {len(final_df)}")
        return len(final_df)
    except Exception as e:
        print(f"❌ overwrite_market_data_with_candles failed: {e}")
        return 0


def rebuild_today_full_session(fetch_api_fn):
    """
    Rebuild today's candles from 09:15 to now
    """

    now = datetime.now()
    market_start = now.replace(hour=9, minute=15, second=0, microsecond=0)

    if now <= market_start:
        print("🟡 Before market open → skip rebuild")
        return

    print(f"🟡 Rebuilding full session: {market_start} → {now}")

    candles = fetch_api_fn(market_start, now)

    if not candles:
        print("❌ No candles returned for full rebuild")
        return

    print(f"✅ Full session candles fetched: {len(candles)}")

    overwrite_market_data_with_candles(candles)

def merge_backfill_candles(api_candles):
    """
    Kept for compatibility with strategy.py.
    Internally now uses overwrite-style rebuild.
    """
    if not api_candles:
        return

    print(f"🔧 Running overwrite merge for {len(api_candles)} candles")
    overwrite_market_data_with_candles(api_candles)


def detect_candle_gap():
    """
    FINAL VERSION:
    - Same day → gap repair
    - New day → skip gap (handled by rebuild logic)
    """

    last_time = get_last_saved_candle_time()
    print(f"DEBUG last_saved_candle_time = {last_time}")

    if last_time is None:
        return False, None, None

    now = datetime.now()

    # ✅ NEW FIX → detect new day
    if last_time.date() != now.date():
        print("🟡 New trading day detected → skip gap detection")
        return False, None, None

    # 🕒 Market timings
    market_start = now.replace(hour=9, minute=15, second=0, microsecond=0)
    market_end = now.replace(hour=15, minute=29, second=0, microsecond=0)

    if now < market_start:
        return False, None, None

    elif now > market_end:
        last_complete = market_end
    else:
        last_complete = now.replace(second=0, microsecond=0) - pd.Timedelta(minutes=1)

    expected_next = last_time + pd.Timedelta(minutes=1)

    if last_time >= market_end:
        return False, None, None

    if expected_next <= last_complete:
        print(f"⚠️ Same-day candle gap detected: {expected_next} → {last_complete}")
        return True, expected_next, last_complete

    return False, None, None

def restore_market_data_clean():
    """
    Final clean restore:
    - yesterday candles → hidden buffer
    - today candles → visible session
    - indicators recomputed immediately
    """

    try:
        df = pd.read_csv(MARKET_DATA_FILE)
    except Exception:
        print("No market history available")
        STATE.history_buffer.clear()
        STATE.session_candles = []
        return

    if df.empty:
        STATE.history_buffer.clear()
        STATE.session_candles = []
        return

    # Parse time properly
    df["time"] = _parse_mixed_market_time(df["time"])
    df = df.dropna(subset=["time"]).sort_values("time")

    today = pd.Timestamp.now().date()

    # Split
    today_df = df[df["time"].dt.date == today]
    prev_df = df[df["time"].dt.date < today]

    # Take last 120 candles for warm-up
    warmup = prev_df.tail(120)

    hidden = pd.concat([warmup, today_df])

    # Fill hidden buffer
    STATE.history_buffer.clear()
    for _, row in hidden.iterrows():
        STATE.history_buffer.append(row.to_dict())

    # Visible only today
    STATE.session_candles = today_df.to_dict("records")

    print(f"📦 Hidden candles: {len(STATE.history_buffer)}")
    print(f"📊 Today candles: {len(STATE.session_candles)}")

    # recompute indicators immediately
    from app.indicators import recompute_indicators_full
    recompute_indicators_full()
def rebuild_state_from_trade_log():
    from app.config import TRADE_LOG_FILE
    from openpyxl import load_workbook

    if not TRADE_LOG_FILE.exists():
        return None

    try:
        wb = load_workbook(TRADE_LOG_FILE)
        ws = wb["Trades"]

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        if not rows:
            return None

        # filter today trades
        today = datetime.now().date()

        valid_rows = []
        for r in rows:
            try:
                entry_time = datetime.strptime(str(r[1]), "%Y-%m-%d %H:%M:%S")
                if entry_time.date() == today:
                    valid_rows.append(r)
            except:
                continue

        if not valid_rows:
            return None

        headers = [cell.value for cell in ws[1]]
        tc_index = headers.index("TradeCount")
        
        trade_counts = [
            int(row[tc_index]) for row in valid_rows
            if row[tc_index] is not None
        ]
        trade_count = max(trade_counts) if trade_counts else 0
        print(f"{trade_count}, ******No.of Trades******" )
      
        realized_pnl = 0.0
        open_position = None
        last_exit_time = None

        for r in valid_rows:
            status = str(r[13]).upper() if r[13] else ""

            if status == "CLOSED":
                try:
                    realized_pnl += float(r[9] or 0)
                    last_exit_time = r[2]
                except:
                    pass

            elif status == "OPEN":
                open_position = {
                    "symbol": r[3],
                    "instrument": r[4],
                    "side": r[5],
                    "remaining_qty": int(r[6]),
                    "entry_price": float(r[7]),
                    "sl": float(r[16]) if r[16] else None,
                    "target": float(r[17]) if r[17] else None,
                    "status": "OPEN",
                    "entry_time": str(r[1]),
                    "paper_trade": True,
                    "entry_logged": True,
                    "trade_id": r[0],
                    "entry_reason": r[18],
                }

        return {
            "trade_count": trade_count,
            "realized_pnl": realized_pnl,
            "current_position": open_position,
            "last_exit_time": last_exit_time,
        }

    except Exception as e:
        print("❌ Trade log rebuild failed:", e)
        return None